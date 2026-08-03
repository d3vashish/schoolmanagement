from datetime import datetime, timedelta, timezone

from fastapi import APIRouter, Depends, HTTPException, Query, status
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.database import get_db
from app.core.deps import role_required
from app.core.security import hash_password
from app.modules.admin.models import SchoolSetting
from app.modules.admin.schemas import (
    AdminDashboardStats,
    AuditLogResponse,
    PaginatedResponse,
    ResetPasswordRequest,
    SettingResponse,
    SettingUpdate,
    UserCreateAdmin,
    UserResponseAdmin,
    UserUpdateAdmin,
)
from app.modules.audit.models import AuditLog
from app.modules.auth.models import ROLES, User

router = APIRouter(prefix="/admin", tags=["admin"])


async def _get_current_super_admin(
    # principal is treated as an Administrator across the app (role label,
    # frontend roleConfig = '*', RoleGuard allows /admin), so admin endpoints
    # accept principal too — otherwise the Admin UI renders but every data call 403s.
    current_user: dict = role_required("super_admin", "principal"),
) -> dict:
    return current_user


@router.get("/dashboard", response_model=AdminDashboardStats)
async def get_admin_dashboard(
    db: AsyncSession = Depends(get_db),
    _: dict = Depends(_get_current_super_admin),
):
    total_users_result = await db.execute(
        select(func.count(User.id)).where(User.deleted_at.is_(None))
    )
    total_users = total_users_result.scalar() or 0

    roles_result = await db.execute(
        select(User.role, func.count(User.id)).where(User.deleted_at.is_(None)).group_by(User.role)
    )
    users_by_role = {row[0]: row[1] for row in roles_result.fetchall()}

    cutoff = datetime.now(timezone.utc) - timedelta(hours=24)
    active_result = await db.execute(
        select(func.count(AuditLog.changed_by.distinct()))
        .where(AuditLog.created_at >= cutoff)
    )
    active_today = active_result.scalar() or 0

    return AdminDashboardStats(
        total_users=total_users,
        users_by_role=users_by_role,
        active_today=active_today,
        storage_used_mb=0.0,
    )


@router.get("/users", response_model=PaginatedResponse)
async def list_users_admin(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    role: str | None = None,
    status: bool | None = None,
    search: str | None = None,
    db: AsyncSession = Depends(get_db),
    _: dict = Depends(_get_current_super_admin),
):
    q = select(User).where(User.deleted_at.is_(None))
    if role:
        q = q.where(User.role == role)
    if status is not None:
        q = q.where(User.is_active == status)
    if search:
        q = q.where(User.email.ilike(f"%{search}%"))
    count_q = select(func.count()).select_from(q.subquery())
    total = (await db.execute(count_q)).scalar() or 0
    q = q.order_by(User.created_at.desc()).offset((page - 1) * page_size).limit(page_size)
    result = await db.execute(q.options(
        selectinload(User.student_profiles),
        selectinload(User.staff_profiles),
        selectinload(User.parent_profiles),
    ))
    users = result.scalars().all()
    data = []
    for u in users:
        profile = None
        if u.role == "student" and u.student_profiles:
            profile = u.student_profiles[0]
        elif u.role in ("teacher", "accountant", "librarian") and u.staff_profiles:
            profile = u.staff_profiles[0]
        elif u.role == "parent" and u.parent_profiles:
            profile = u.parent_profiles[0]
        data.append(UserResponseAdmin(
            id=u.id,
            email=u.email,
            phone=u.phone,
            role=u.role,
            is_active=u.is_active,
            first_name=getattr(profile, "first_name", None),
            last_name=getattr(profile, "last_name", None),
            created_at=u.created_at,
        ))
    return PaginatedResponse(data=[d.model_dump() for d in data], total=total, page=page, page_size=page_size)


@router.post("/users", response_model=UserResponseAdmin, status_code=status.HTTP_201_CREATED)
async def create_user_admin(
    body: UserCreateAdmin,
    db: AsyncSession = Depends(get_db),
    _: dict = Depends(_get_current_super_admin),
):
    existing = await db.execute(select(User).where(User.email == body.email, User.deleted_at.is_(None)))
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Email already registered")
    if body.role not in ROLES:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Invalid role. Must be one of: {', '.join(ROLES)}")
    user = User(
        email=body.email,
        phone=body.phone,
        hashed_pw=hash_password(body.password),
        role=body.role,
        is_active=body.is_active,
    )
    db.add(user)
    await db.flush()
    await db.refresh(user)
    return UserResponseAdmin(
        id=user.id,
        email=user.email,
        phone=user.phone,
        role=user.role,
        is_active=user.is_active,
        first_name=body.first_name,
        last_name=body.last_name,
        created_at=user.created_at,
    )


@router.get("/users/{user_id}", response_model=UserResponseAdmin)
async def get_user_admin(
    user_id: str,
    db: AsyncSession = Depends(get_db),
    _: dict = Depends(_get_current_super_admin),
):
    user = await db.get(User, user_id, options=[
        selectinload(User.student_profiles),
        selectinload(User.staff_profiles),
        selectinload(User.parent_profiles),
    ])
    if not user or user.deleted_at is not None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    profile = None
    if user.role == "student" and user.student_profiles:
        profile = user.student_profiles[0]
    elif user.role in ("teacher", "accountant", "librarian") and user.staff_profiles:
        profile = user.staff_profiles[0]
    elif user.role == "parent" and user.parent_profiles:
        profile = user.parent_profiles[0]
    return UserResponseAdmin(
        id=user.id,
        email=user.email,
        phone=user.phone,
        role=user.role,
        is_active=user.is_active,
        first_name=getattr(profile, "first_name", None),
        last_name=getattr(profile, "last_name", None),
        created_at=user.created_at,
    )


@router.put("/users/{user_id}", response_model=UserResponseAdmin)
async def update_user_admin(
    user_id: str,
    body: UserUpdateAdmin,
    db: AsyncSession = Depends(get_db),
    _: dict = Depends(_get_current_super_admin),
):
    user = await db.get(User, user_id)
    if not user or user.deleted_at is not None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    if body.email is not None:
        user.email = body.email
    if body.phone is not None:
        user.phone = body.phone
    if body.role is not None:
        if body.role not in ROLES:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Invalid role. Must be one of: {', '.join(ROLES)}")
        user.role = body.role
    if body.is_active is not None:
        user.is_active = body.is_active
    await db.flush()
    await db.refresh(user, [
        "student_profiles",
        "staff_profiles",
        "parent_profiles",
    ])
    return UserResponseAdmin(
        id=user.id,
        email=user.email,
        phone=user.phone,
        role=user.role,
        is_active=user.is_active,
        first_name=body.first_name,
        last_name=body.last_name,
        created_at=user.created_at,
    )


@router.patch("/users/{user_id}/status", response_model=UserResponseAdmin)
async def toggle_user_status(
    user_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(_get_current_super_admin),
):
    if user_id == current_user["id"]:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Cannot disable your own account")
    user = await db.get(User, user_id)
    if not user or user.deleted_at is not None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    user.is_active = not user.is_active
    await db.flush()
    return UserResponseAdmin(
        id=user.id,
        email=user.email,
        phone=user.phone,
        role=user.role,
        is_active=user.is_active,
        created_at=user.created_at,
    )


@router.post("/users/{user_id}/reset-password", status_code=status.HTTP_200_OK)
async def reset_user_password(
    user_id: str,
    body: ResetPasswordRequest,
    db: AsyncSession = Depends(get_db),
    _: dict = Depends(_get_current_super_admin),
):
    user = await db.get(User, user_id)
    if not user or user.deleted_at is not None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    user.hashed_pw = hash_password(body.new_password)
    await db.flush()
    return {"message": "Password reset successful"}


@router.get("/settings", response_model=list[SettingResponse])
async def get_settings(
    db: AsyncSession = Depends(get_db),
    _: dict = Depends(_get_current_super_admin),
):
    result = await db.execute(select(SchoolSetting).order_by(SchoolSetting.key))
    settings = result.scalars().all()
    return [SettingResponse(key=s.key, value=s.value, updated_at=s.updated_at) for s in settings]


@router.put("/settings", response_model=list[SettingResponse])
async def update_settings(
    body: list[SettingUpdate],
    db: AsyncSession = Depends(get_db),
    _: dict = Depends(_get_current_super_admin),
):
    results = []
    for item in body:
        existing = await db.execute(
            select(SchoolSetting).where(SchoolSetting.key == item.key)
        )
        setting = existing.scalar_one_or_none()
        if setting:
            setting.value = item.value
        else:
            setting = SchoolSetting(key=item.key, value=item.value)
            db.add(setting)
        results.append(setting)
    await db.flush()
    for s in results:
        await db.refresh(s)
    return [SettingResponse(key=s.key, value=s.value, updated_at=s.updated_at) for s in results]


@router.get("/audit/log", response_model=PaginatedResponse)
async def get_audit_log(
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    action_type: str | None = None,
    user_id: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    db: AsyncSession = Depends(get_db),
    _: dict = Depends(_get_current_super_admin),
):
    q = select(AuditLog)
    if action_type:
        q = q.where(AuditLog.action == action_type)
    if user_id:
        q = q.where(AuditLog.changed_by == user_id)
    if date_from:
        q = q.where(AuditLog.created_at >= datetime.fromisoformat(date_from))
    if date_to:
        q = q.where(AuditLog.created_at <= datetime.fromisoformat(date_to))
    count_q = select(func.count()).select_from(q.subquery())
    total = (await db.execute(count_q)).scalar() or 0
    q = q.order_by(AuditLog.created_at.desc()).offset((page - 1) * page_size).limit(page_size)
    result = await db.execute(q)
    logs = result.scalars().all()
    data = []
    for log in logs:
        data.append(AuditLogResponse(
            id=log.id,
            table_name=log.table_name,
            record_id=log.record_id,
            action=log.action,
            changed_by=log.changed_by,
            summary=f"{log.action} {log.table_name}",
            created_at=log.created_at,
        ))
    return PaginatedResponse(data=[d.model_dump() for d in data], total=total, page=page, page_size=page_size)

@router.post("/backup-now")
async def backup_now(
    db: AsyncSession = Depends(get_db),
    _: dict = Depends(_get_current_super_admin),
):
    """On-demand backup. Writes a restorable SQL dump AND a readable Excel workbook
    (one sheet per table), saving both to local app data and uploading to the cloud."""
    import os
    import asyncio
    import subprocess
    import datetime as _dt
    import uuid as _uuid
    from decimal import Decimal
    from pathlib import Path
    from sqlalchemy import text

    container = os.getenv("DB_CONTAINER", "school-erp-backend-db-1")
    db_user = os.getenv("DB_USER_BACKUP", "erp_user")
    db_name = os.getenv("DB_NAME_BACKUP", "school_erp")

    stamp = datetime.now(timezone.utc).strftime("%Y-%m-%d_%H%M%S")
    base = os.getenv("LOCALAPPDATA") or os.path.expanduser("~")
    local_dir = Path(base) / "OCC-Backups"
    local_dir.mkdir(parents=True, exist_ok=True)

    sql_name = f"school_erp_{stamp}.sql"
    xlsx_name = f"school_erp_{stamp}.xlsx"
    sql_path = local_dir / sql_name
    xlsx_path = local_dir / xlsx_name

    # ---------- 1) SQL dump (restorable) ----------
    def _dump():
        with open(sql_path, "wb") as f:
            proc = subprocess.run(
                ["docker", "exec", container, "pg_dump", "-U", db_user, db_name],
                stdout=f, stderr=subprocess.PIPE,
            )
        return proc.returncode, proc.stderr.decode(errors="ignore")

    sql_ok = False
    try:
        rc, err = await asyncio.to_thread(_dump)
        sql_ok = (rc == 0 and sql_path.exists() and sql_path.stat().st_size > 0)
    except FileNotFoundError:
        err = "docker not found on server PATH"

    # ---------- 2) Excel workbook (readable) ----------
    tables = (await db.execute(text(
        "SELECT tablename FROM pg_tables WHERE schemaname='public' ORDER BY tablename"
    ))).scalars().all()

    table_data = {}
    for tbl in tables:
        res = await db.execute(text(f'SELECT * FROM "{tbl}"'))
        cols = list(res.keys())
        rows = res.fetchall()
        table_data[tbl] = (cols, [list(r) for r in rows])

    def _cell(v):
        if v is None:
            return ""
        if isinstance(v, (_dt.datetime, _dt.date)):
            return v.isoformat()
        if isinstance(v, _uuid.UUID):
            return str(v)
        if isinstance(v, Decimal):
            return float(v)
        if isinstance(v, (bytes, bytearray)):
            return "<binary>"
        return v

    def _build_xlsx():
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        wb = Workbook()
        wb.remove(wb.active)
        used = set()
        for tbl, (cols, rows) in table_data.items():
            name = tbl[:31] or "sheet"
            base_name = name
            i = 1
            while name in used:
                suffix = f"_{i}"
                name = base_name[:31 - len(suffix)] + suffix
                i += 1
            used.add(name)
            ws = wb.create_sheet(title=name)
            ws.append(cols)
            for c in ws[1]:
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="15803D")
            for row in rows:
                ws.append([_cell(v) for v in row])
        if not wb.sheetnames:
            wb.create_sheet(title="empty")
        wb.save(str(xlsx_path))

    xlsx_ok = False
    xlsx_err = None
    try:
        await asyncio.to_thread(_build_xlsx)
        xlsx_ok = xlsx_path.exists() and xlsx_path.stat().st_size > 0
    except ModuleNotFoundError:
        xlsx_err = "openpyxl not installed on the server (run: pip install openpyxl)"
    except Exception as e:  # noqa: BLE001
        xlsx_err = str(e)[:200]

    if not sql_ok and not xlsx_ok:
        raise HTTPException(status_code=500, detail=f"Backup failed. SQL: {err[:200] if 'err' in dir() else 'n/a'}. Excel: {xlsx_err}")

    # ---------- 3) Upload both to cloud (best-effort) ----------
    cloud_msgs = []
    try:
        from app.shared.storage import upload_file
        if sql_ok:
            await upload_file(str(sql_path), f"backups/{sql_name}")
            cloud_msgs.append("SQL")
        if xlsx_ok:
            await upload_file(str(xlsx_path), f"backups/{xlsx_name}")
            cloud_msgs.append("Excel")
        cloud = ("uploaded to cloud: " + ", ".join(cloud_msgs)) if cloud_msgs else "nothing uploaded"
    except Exception as e:  # noqa: BLE001
        cloud = f"local only (cloud upload failed: {str(e)[:150]})"

    return {
        "message": "Backup complete",
        "sql_file": sql_name if sql_ok else None,
        "excel_file": xlsx_name if xlsx_ok else None,
        "excel_error": xlsx_err,
        "local_dir": str(local_dir),
        "cloud": cloud,
        "tables": len(table_data),
    }